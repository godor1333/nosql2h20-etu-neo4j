import os

from flask_restful import Resource
from flask import send_file, request
from neomodel import db
from openpyxl import Workbook

from uemployees.faculty.models import Faculty
from uemployees.department.models import Department
from uemployees.employee.models import Employee
from uemployees.degree.models import Degree
from uemployees.discipline.models import Discipline
from uemployees.contrib.excel_parser import csvFromExcelParser
from uemployees.db import load_db_data


def createFacultySheet(wb):
    facultySheet = wb.create_sheet('faculty')
    facultySheet['A1'] = 'id'
    facultySheet['B1'] = 'name'

    i = 2
    for faculty in Faculty.nodes.all():
        facultySheet['A{}'.format(i)] = faculty.id
        facultySheet['B{}'.format(i)] = faculty.name
        i += 1


def createDepartmentSheet(wb):
    departmentSheet = wb.create_sheet('department')
    departmentSheet['A1'] = 'id'
    departmentSheet['B1'] = 'name'
    departmentSheet['C1'] = "id_faculty"

    i = 2
    for faculty in Faculty.nodes.all():
        for department in faculty.departments.all():
            departmentSheet['A{}'.format(i)] = department.id
            departmentSheet['B{}'.format(i)] = department.name
            departmentSheet['C{}'.format(i)] = faculty.id
            i += 1


def createEmployeeSheet(wb):
    employeeSheet = wb.create_sheet('employee')
    employeeSheet['A1'] = 'id'
    employeeSheet['B1'] = 'name'
    employeeSheet['C1'] = 'photo_url'
    employeeSheet['D1'] = 'email'
    employeeSheet['E1'] = 'education'

    i = 2
    for employee in Employee.nodes.all():
        employeeSheet['A{}'.format(i)] = employee.id
        employeeSheet['B{}'.format(i)] = employee.name
        employeeSheet['C{}'.format(i)] = employee.photo_url
        employeeSheet['D{}'.format(i)] = employee.email
        employeeSheet['E{}'.format(i)] = employee.education
        i += 1


def createEmployeeDepartmentSheet(wb):
    employeeDepartmentSheet = wb.create_sheet('employee_department')
    employeeDepartmentSheet['A1'] = 'id_department'
    employeeDepartmentSheet['B1'] = 'id_employee'
    employeeDepartmentSheet['C1'] = 'job_title'

    i = 2
    for department in Department.nodes.all():
        for employee in department.employees.all():
            employeeDepartmentSheet['A{}'.format(i)] = department.id
            employeeDepartmentSheet['B{}'.format(i)] = employee.id
            employeeDepartmentSheet['C{}'.format(i)] = department.employees.relationship(employee).job_title
            i += 1


def createPublicationSheet(wb):
    publicationSheet = wb.create_sheet('publication')
    publicationSheet['A1'] = 'id'
    publicationSheet['B1'] = 'content'
    publicationSheet['C1'] = 'id_employee'

    i = 2
    for employee in Employee.nodes.all():
        for publication in employee.publications.all():
            publicationSheet['A{}'.format(i)] = publication.id
            publicationSheet['B{}'.format(i)] = publication.content
            publicationSheet['C{}'.format(i)] = employee.id
            i += 1


def createDegreeSheet(wb):
    degreeSheet = wb.create_sheet('degree')
    degreeSheet['A1'] = 'id'
    degreeSheet['B1'] = 'content'

    i = 2
    for degree in Degree.nodes.all():
        degreeSheet['A{}'.format(i)] = degree.id
        degreeSheet['B{}'.format(i)] = degree.content
        i += 1


def createEmployeeDegreeSheet(wb):
    employeeDegreeSheet = wb.create_sheet('employee_degree')
    employeeDegreeSheet['A1'] = 'id_employee'
    employeeDegreeSheet['B1'] = 'id_degree'

    i = 2
    for employee in Employee.nodes.all():
        for degree in employee.degrees.all():
            employeeDegreeSheet['A{}'.format(i)] = employee.id
            employeeDegreeSheet['B{}'.format(i)] = degree.id
            i += 1


def createDisciplineSheet(wb):
    disciplineSheet = wb.create_sheet('discipline')
    disciplineSheet['A1'] = 'id'
    disciplineSheet['B1'] = 'name'

    i = 2
    for discipline in Discipline.nodes.all():
        disciplineSheet['A{}'.format(i)] = discipline.id
        disciplineSheet['B{}'.format(i)] = discipline.name
        i += 1


def getLessons(emp_id, dis_id):
    lessons = [
        lesson[0].items()
        for lesson in db.cypher_query(
            f'MATCH (e:Employee)'
            f'-[r:TEACH_A_DISCIPLINE]->(d:Discipline) '
            f'WHERE id(e)={emp_id} AND id(d)={dis_id} '
            f'RETURN r;'
        )[0]
    ]

    return [{
        k: v for k, v in lesson
    } for lesson in lessons]


def getUniqueDisciplines(desciplines):
    unique_disciplines = []
    for discipline in desciplines:
        if discipline not in unique_disciplines:
            unique_disciplines.append(discipline)
    return unique_disciplines


def createEmployeeDisciplineSheet(wb):
    employeeDisciplineSheet = wb.create_sheet('employee_discipline')
    employeeDisciplineSheet['A1'] = 'id_employee'
    employeeDisciplineSheet['B1'] = 'id_discipline'
    employeeDisciplineSheet['C1'] = 'group'
    employeeDisciplineSheet['D1'] = 'time'
    employeeDisciplineSheet['E1'] = 'auditorium'

    i = 2
    for employee in Employee.nodes.all():
        for discipline in getUniqueDisciplines(employee.disciplines.all()):
            for lesson in getLessons(employee.id, discipline.id):
                employeeDisciplineSheet['A{}'.format(i)] = employee.id
                employeeDisciplineSheet['B{}'.format(i)] = discipline.id

                employeeDisciplineSheet['C{}'.format(i)] = lesson['group']
                employeeDisciplineSheet['D{}'.format(i)] = lesson['time']
                employeeDisciplineSheet['E{}'.format(i)] = lesson['auditorium']
                i += 1


def createExportFile(file_name):
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(THIS_FOLDER, file_name)

    wb = Workbook()

    createFacultySheet(wb)
    createDepartmentSheet(wb)
    createEmployeeSheet(wb)
    createEmployeeDepartmentSheet(wb)
    createPublicationSheet(wb)
    createDegreeSheet(wb)
    createEmployeeDegreeSheet(wb)
    createDisciplineSheet(wb)
    createEmployeeDisciplineSheet(wb)

    wb.save(file_path)

    return file_path


class ExportView(Resource):
    def get(self):
        file_name = 'document_export.xlsx'
        file_path = createExportFile(file_name)

        response = send_file(
            file_path,
            as_attachment=True,
            attachment_filename=file_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            cache_timeout=0
        )

        response.headers["Content-Length"] = os.path.getsize(file_path)
        response.headers["Content-Disposition"] = "attachment; filename={}".format(file_name)
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return response


class ImportView(Resource):
    def post(self):
        try:
            file = request.files['import_document']
            if file.filename.split('.')[-1] != 'xlsx':
                return '', 400
            file_dst = os.path.abspath(f'tmp/import_document.xlsx')
            file.save(file_dst)
            csvFromExcelParser(file_dst)
            load_db_data()
            return 'success', 200
        except:
            return '', 400


class GetCSVView(Resource):
    def get(self, file_name):
        file_path = f'tmp/{file_name}'

        response = send_file(
            file_path,
            as_attachment=True,
            attachment_filename=file_name,
            mimetype='application/csv',
            cache_timeout=0
        )

        response.headers["Content-Length"] = os.path.getsize(file_path)
        response.headers["Content-Type"] = "application/csv"

        return response
